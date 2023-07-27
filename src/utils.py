import sys
import os
import win32com.client as win32
from pathlib import Path
import datetime
from openpyxl import load_workbook
from openpyxl.styles import numbers
import pyexcel
import openpyxl
import json
import pandas as pd
import re

class pathsProyect:
    def __init__(self) -> None:
        self.jsonCcaj = None
        self.jsonCcob = None
        self.jsonCashOut = None
        self.csvCashOut = None
        self.dirCcaj = None
        self.dirCcob = None
        self.dirCashOut = None
        self.appPath=None
        self.folderProyect=None
        self.bot1=None
        self.bot1_config=None
        self.tables=None
        self.jsonClientBox=None
        self.jsonCobBox=None
        self.detalleCajaCsv=None
        self.billsTableCsv=None
        self.testMode=False
        self.get_app_path()
        self.getting_paths()
    def get_app_path(self):
        if getattr(sys, 'frozen', False):
            application_path = os.path.dirname(sys.executable)
        elif __file__:
            application_path = os.path.dirname(__file__)
        self.appPath =Path(application_path)
        self.folderProyect=self.appPath.parent.absolute()
        #return Path(application_path)
    def getting_paths(self):
        self.jsonCcaj=os.path.join(self.appPath.parent.absolute(),"src","target","CashClosingInfo.json")
        self.jsonCashOut=os.path.join(self.appPath.parent.absolute(),"src","target","CashOutInfo.json")
        self.jsonCcob=os.path.join(self.appPath.parent.absolute(),"src","target","CollectorClosingFilesDonwload.json")
        self.bot1=os.path.join(self.appPath.parent.absolute().parent.absolute(),"Bot1","SapHunter")
        self.bot1_extractos=os.path.join(self.bot1,"extractosBancarios")
        self.bot1_plantillas=os.path.join(self.bot1,"plantillasSap")
        self.bot1_config=os.path.join(self.bot1,"config.xlsx")
        self.tables=os.path.join(self.appPath.parent.absolute(),"Tablas")
        self.dirCcaj=os.path.join(self.appPath.parent.absolute(),"Cierres de Caja","formatoxlsx")
        self.csvCashOut=os.path.join(self.appPath.parent.absolute(),"Tablas","cashOut.csv")
        self.detalleCajaCsv=os.path.join(self.appPath.parent.absolute(),"Tablas","DetalleCcajTable.csv")
        self.billsTableCsv=os.path.join(self.appPath.parent.absolute(),"Tablas","billsTable.csv")
        self.coinsTableCsv=os.path.join(self.appPath.parent.absolute(),"Tablas","coinsTable.csv")
        self.bankTransferTableCsv=os.path.join(self.appPath.parent.absolute(),"Tablas","banktransfersTable.csv")
        self.vouchersTableCsv=os.path.join(self.appPath.parent.absolute(),"Tablas","voucherTable.csv")
        self.checksTableCsv=os.path.join(self.appPath.parent.absolute(),"Tablas","checksTable.csv")
        self.qrTableCsv=os.path.join(self.appPath.parent.absolute(),"Tablas","qrTable.csv'")
        self.cuoponsTableCsv=os.path.join(self.appPath.parent.absolute(),"Tablas","cuoponTable.csv")
        self.jsonCobBox=os.path.join(self.appPath.parent.absolute(),"src","target","CcobBox.json")
        self.jsonClientBox=os.path.join(self.appPath.parent.absolute(),"src","target","CcajBox.json")
        self.jsonFinal=os.path.join(self.appPath.parent.absolute(),"src","target","FinalDataToExcel.json")
        self.FullExcelDataJson=os.path.join(self.appPath.parent.absolute(),"src","target","FullExcelData.json")
paths=pathsProyect()
def get_current_path():
    config_name = 'myapp.cfg'
    # determine if application is a script file or frozen exe
    if getattr(sys, 'frozen', False):
        application_path = os.path.dirname(sys.executable)
    elif __file__:
        application_path = os.path.dirname(__file__)
    application_path2 = Path(application_path)
    return application_path2.parent.absolute()
def get_tables_path():
    config_name = 'myapp.cfg'
    # determine if application is a script file or frozen exe
    if getattr(sys, 'frozen', False):
        application_path = os.path.dirname(sys.executable)
    elif __file__:
        application_path = os.path.dirname(__file__)
    application_path2 = Path(application_path)
    return os.path.join(application_path2.parent.absolute(),"Tablas")


def remove_files(folderPath):
    for path in os.listdir(folderPath):
    # check if current path is a file
        if os.path.isfile(os.path.join(folderPath, path)):
            if path[-4:]==".xls" or path[-5:]==".xlsx" or path[-4:]==".csv" or path[-4:]==".txt" :
                #if path=="auszug.txt" or path=="umsatz.txt":
                os.remove(os.path.join(folderPath, path))
                #print("txt file deleted")
def delete_xlsFiles():
    print("Borrando archivos anteriores...")
    paths=pathsProyect()
    data={"data":[]}
    with open(paths.jsonCcaj,"w") as json_file:
        json_file.write(json.dumps(data,indent=4))
    with open(paths.jsonCcob,"w") as json_file:
        json_file.write(json.dumps(data,indent=4))
    remove_files(os.path.join(paths.folderProyect,"Cierres de Caja"))
    remove_files(os.path.join(paths.folderProyect,"Cierres de Caja","formatoxlsx"))
    remove_files(os.path.join(paths.folderProyect,"Cierres de Cobrador"))
    remove_files(os.path.join(paths.folderProyect,"Cierres de Cobrador","formatoxlsx"))
    remove_files(os.path.join(paths.folderProyect,"SapInfo"))
    remove_files(os.path.join(paths.folderProyect,"Ouputs"))
def convert_xls(pathFolder):    
    filesInfolder=os.listdir(pathFolder)
    e=""
    for file in filesInfolder:
        if file.endswith(".xls"):
            try:
                #print(file)
                #name of file 
                xls=os.path.join(pathFolder,file)
                xlsx=os.path.join(pathFolder,"formatoxlsx",file.replace(".xls",".xlsx"))
                pyexcel.save_book_as(file_name=xls, dest_file_name=xlsx)
            except Exception as e:
                print(e)
                os.remove(xlsx)
    return e

def get_index_columns_config():
    tableJsonConfig=os.path.join(get_current_path(),"src","target","indexColumnsConfig.json")
    with open(tableJsonConfig) as json_file:
        data = json.load(json_file)
    return data
def get_kwords_rowLimits_config():
    tableJsonConfig=os.path.join(get_current_path(),"src","target","kwordsRowLimitsConfig.json")
    with open(tableJsonConfig) as json_file:
        data = json.load(json_file)
    return data
def get_currency(fileName):
    #print(fileName)
    path=Path(fileName)
    fileName=path.name
    #get the fileName in path 

    if fileName.find("Us")!=-1:
        typeCurrency="dólar"
    elif fileName.find("Bs")!=-1:
        typeCurrency="Bs"
    elif fileName.find("First")!=-1:
        typeCurrency="ambos"
    else:
        typeCurrency="other"
    return typeCurrency

def writeJson():
    with open(paths.jsonCcaj,"r") as json_file:
        data = json.load(json_file)
    for row in data['data']:
        row['NuevaData']={}
    with open(paths.jsonCcaj,"w") as json_file:
        json.dump(data,json_file,indent=4)
def getSgvData(fileName):
    code=re.findall(r'(\d{5})_', fileName)[0]
    with open(paths.jsonCcaj,"r") as json_file:
        data = json.load(json_file)
    for d in data["data"]:
        if d['Código']==code:
            return d
def normalizeTable():
    #print("Normalizando tablas")
    dfs=[]
    dfNames=[paths.billsTableCsv,paths.coinsTableCsv,paths.bankTransferTableCsv
             ,paths.vouchersTableCsv,paths.checksTableCsv,paths.qrTableCsv
             ,paths.cuoponsTableCsv]
    
    for dfName in dfNames:
        try:
            df=pd.read_csv(dfName,sep=';')
            dfs.append(df)
        except:
            pass
            print("-------------No se encontro data en el archivo: ",dfName)
    try:
        df_all=pd.concat(dfs,ignore_index=True)
        allData=df_all.to_dict('records')
    except:
        print("-------------No se encontro data en ningun archivo")
        allData=[]
    df_all=pd.DataFrame(allData)
    df_all.to_csv(paths.detalleCajaCsv,index=False,sep=';',header=True)

def loginInfo():
    wb=openpyxl.load_workbook(os.path.join(get_current_path(),"config.xlsx"))
    configDat={}
    ws=wb["login"]
    configDat['dates']={}
    dinit=ws["B2"].value
    dinit=dinit+datetime.timedelta(days=1)
    configDat['dates']['dInit']=dinit
    configDat['dates']['dEnd']=ws["B3"].value
    configDat['users']={}
    configDat['SapLogin']={
                'SAPPath': ws['B16'].value,
                'user': ws['B17'].value,
                'psw': ws['B18'].value,
                'environment': ws['B19'].value,
                'layout': ws['B20'].value,
                'fechaInicio': ws['B2'].value,
                'fechaFin': ws['B3'].value
                }
    listOfAccounts=[]
    wsAccounts = wb['ctasMayores']
    for i in range(2, wsAccounts.max_row+1):
        accountCell = wsAccounts[f'A{i}'].value
        accountCell = str(accountCell)
        accountCell = accountCell.replace(" ", "")
        if accountCell != None and accountCell != "":
            listOfAccounts.append(accountCell)
    
    configDat['accounts'] = listOfAccounts
    maxRow=ws.max_row
    for i in range(2,maxRow+1):
        if ws["G"+str(i)].value!=None and ws["F"+str(i)].value=="SI":
            configDat['users'][ws["G"+str(i)].value]={}
            configDat['users'][ws["G"+str(i)].value]['user']=ws["H"+str(i)].value
            configDat['users'][ws["G"+str(i)].value]['password']=ws["I"+str(i)].value
            configDat['users'][ws["G"+str(i)].value]['recaudadora']=ws["J"+str(i)].value


    configDat['flags']={}
    configDat['flags']['flow']=ws["B5"].value
    configDat['flags']['cumulative']=ws["B6"].value
    return configDat

def configToJson():
    configxlsxPath=os.path.join(get_current_path(),"config.xlsx")
    indexColumnsPathJson=os.path.join(get_current_path(),"src","target","indexColumnsConfig.json")
    kwordsRowLimitsPathJson=os.path.join(get_current_path(),"src","target","kwordsRowLimitsConfig.json")

    dfC=pd.read_excel(configxlsxPath,sheet_name="columnas")
    #print(df)
    #conver the df into collection of dictionaries
    dataColumns=dfC.values.tolist()
    columnsDict = {}
    for d in dataColumns:
        if d[0] not in columnsDict:
            columnsDict[d[0]] = {}
        if d[1] not in columnsDict[d[0]]:
            columnsDict[d[0]][d[1]] = {}
        if d[2] not in columnsDict[d[0]][d[1]]:
            columnsDict[d[0]][d[1]][d[2]] = {}
        if d[3] not in columnsDict[d[0]][d[1]][d[2]]:
            columnsDict[d[0]][d[1]][d[2]][d[3]] = {}
        if d[4] not in columnsDict[d[0]][d[1]][d[2]][d[3]]:
            columnsDict[d[0]][d[1]][d[2]][d[3]][d[4]] = {}
        columnsDict[d[0]][d[1]][d[2]][d[3]][d[4]][d[5]] = d[6]
    with open(indexColumnsPathJson, 'w') as outfile:
        json.dump(columnsDict, outfile,indent=4)


    dfKwords=pd.read_excel(configxlsxPath,sheet_name="kwords")
    dataKeywords=dfKwords.values.tolist()
    kwordsDict = {}
    for d in dataKeywords:
        if d[0] not in kwordsDict:
            kwordsDict[d[0]] = {}
        if d[1] not in kwordsDict[d[0]]:
            kwordsDict[d[0]][d[1]] = {}
        if d[2] not in kwordsDict[d[0]][d[1]]:
            kwordsDict[d[0]][d[1]][d[2]] = {}
        kwordsDict[d[0]][d[1]][d[2]][d[3]] = d[4]

    with open(kwordsRowLimitsPathJson, 'w') as outfile:
        json.dump(kwordsDict, outfile,indent=4)

def get_bot1_configData():
    df_bot1=pd.read_excel(paths.bot1_config,sheet_name="cuentas")
    data=df_bot1.to_dict('records')
    newData=[]
    r={}
    for d in data:
        fordigits=str(d['NRO.CUENTA'])[-4:]
        r[fordigits]={}
        r[fordigits]['BANCO']=d['ENTIDAD FINANCIERA']
        r[fordigits]['NRO.CUENTA']=d['NRO.CUENTA']
    return r
def get_templatesSap(dates):
    d1=dates["dInit"]
    #rest 1 day to d1
    d1=d1+datetime.timedelta(days=-5)
    #print(d1)
    d2=dates["dEnd"]
    dirs=[]
    banksData=get_bot1_configData()
    for f in os.scandir(paths.bot1_plantillas):
        if f.is_dir():
            file={
                "name":f.name,
                "path":f.path
            }
            dateName=datetime.datetime.strptime(file['name'],"%d%m%Y")
            if dateName>=d1 and dateName<=d2:
                dirs.append(file)
    files=[]
    extractInfo=[]
    for dir in dirs:
        for f in os.scandir(dir['path']):
            if f.name.endswith(".xlsx"):
                file={
                    "name":f.name,
                    "path":f.path
                }
                files.append(file)
                df=pd.read_excel(file["path"],sheet_name='UNION',header=15)
                values=df.values.tolist()
                fordigits=re.findall(r'(\d{4})-',file['name'])[0]
                for value in values:
                    newvalue={
                        "carpeta":"'"+str(dir['name']),
                        "banco":banksData[fordigits]['BANCO'],
                        "nro cuenta4digits":"'"+str(fordigits),
                        "nro cuenta":"'"+str(banksData[fordigits]['NRO.CUENTA']),
                        "date":value[0],
                        "Nro Documento":"'"+str(value[1]),
                        "Descripcion":str(value[2]).replace("=",""),
                        "importe":value[4],
                    }
                    extractInfo.append(newvalue)
    df_templatesSap=pd.DataFrame(extractInfo)
    df_templatesSap.to_csv(os.path.join(paths.tables,"ExtractosBancarios.csv"),index=False,sep=	";",encoding="utf-8")

def concat_dfs2(lists):
    df_list=[]
    for lista in lists:
        #df is not empty
        df=pd.DataFrame(lista)
        if df.empty==False:
            df_list.append(df)
    dff=pd.concat(df_list,axis=1)
    concat_table=dff.to_dict('records')
    return concat_table
def concat_dfs(lists):
    df_list=[]
    for lista in lists:
        #df is not empty
        df=pd.DataFrame(lista)
        df_list.append(df)
    dff=pd.concat(df_list,axis=1)
    concat_table=dff.to_dict('records')
    return concat_table

def change_ExcelSeparators():
    # Crear una instancia de la aplicación Exce

    # Cargar el archivo de Excel
    archivo_excel = r"C:\DanielBots\bot4\sting.xlsx"
    libro = load_workbook(archivo_excel)

    # Obtener la hoja de trabajo activa
    hoja_activa = libro.active

    hoja_activa.cell(row=1, column=3).value = 1234567890.1234567890
    # Configurar el separador de decimales
    hoja_activa.number_format = numbers.FORMAT_NUMBER_DOT_SEPARATED2

    # Guardar los cambios en el archivo
    libro.save(archivo_excel)

    # Cerrar el libro de Excel
    libro.close()
def change_ExcelSeparators2():
    # Crear una instancia de la aplicación Excel
    excel_app = win32.gencache.EnsureDispatch('Excel.Application')

    # Configurar el separador de decimales
    excel_app.DecimalSeparator = ','

    # Configurar el separador de miles
    excel_app.ThousandsSeparator = '.'

    # Guardar la configuración
    #excel_app.Save()

    # Cerrar la aplicación Excel
    excel_app.Quit()
if __name__ == '__main__':
    #change_ExcelSeparators2()
    print("hola, ejecutandose programa...")

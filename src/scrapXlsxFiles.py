import openpyxl
from utils import get_current_path,get_index_columns_config,get_currency,get_kwords_rowLimits_config,configToJson,get_tables_path
from utils import getSgvData,normalizeTable,pathsProyect
import os
import re
import json
from pathlib import Path
import pandas as pd

paths=pathsProyect()
class scrapTablesExcel:
    def __init__(self,fileName,distributionType) -> None:
        self.XlsxPath=os.path.join(get_current_path(),"Cierres de Caja","formatoxlsx")
        self.indexColumns=get_index_columns_config()
        self.kwordsRowLimits=get_kwords_rowLimits_config()
        self.sh=openpyxl.load_workbook(os.path.join(get_current_path(),"Cierres de Caja","formatoxlsx",fileName)).worksheets[0]
        self.currency=get_currency(fileName)
        self.distributionType=distributionType
        self.sgvData=getSgvData(fileName)
        self.fileName=fileName
        self.codeCcaj=None
        self.currencyNeto=None
        self.get_recauda()
        self.gap=0
    def get_recauda(self):
        filename=self.fileName
        recaud=re.findall(r'(.*)_\d{5}_', filename)[0]
        self.recaud=recaud
        self.CodeCcaj=re.findall(r'_(\d{5})_', filename)[0]
        if self.currency.find("Bs")!=-1:
            self.currencyNeto="Bs"
        else:
            self.currencyNeto=self.currency
    def get_left_up_vertex_table(self,tableName,moneyType):
        sh=self.sh
        i=1
        if self.currency=="dólar":
            pass
        columBill=self.indexColumns[self.distributionType]["Detalle Operaciones"][self.currency][moneyType][tableName]["valor"]
        kword=self.kwordsRowLimits[self.distributionType][self.currency][tableName]["superior"]
        while sh.cell(i,columBill).value !=kword:
            i=i+1
        return i
    def updateCurrency(self):
        sh=self.sh
        j=1
        i=6
        #move to the right until find a number
        while (sh.cell(i,j).value==None):
            j=j+1
        if j==2:
            self.currency="Bs00"
        elif j==4:
            self.currency="Bs"
        else:
            print("Error: no se pudo determinar el gap")
    
    def getBillstable(self):
        sh=self.sh
        filterBillsWords=["Cantidad",None]
        typeCurrency=self.currency
        typeFile=""
        typeDistribution=self.distributionType
        nameTable="billetes"
        columnsTableDict=self.indexColumns[typeDistribution]["Detalle Operaciones"][typeCurrency]["efectivo"][nameTable]
        columBill=columnsTableDict["valor"]
        columQantityBill=columnsTableDict["Cantidad"]
        ColumnAmountBill=columnsTableDict["subtotal"]
        columTotalBill=columnsTableDict["total"]
        billsTable=[]
        downLimitWord=self.kwordsRowLimits[typeDistribution][typeCurrency][nameTable]["inferior"]
        moneyType="efectivo"
        i=self.get_left_up_vertex_table(nameTable,moneyType)
        while sh.cell(i,columTotalBill).value !=downLimitWord:
            billsDict={
                "billValue":sh.cell(i,columBill).value,
                "billQuantity":sh.cell(i,columQantityBill).value
            }
            if sh.cell(i,ColumnAmountBill).value=="-":
                billsDict["AmountBill"]=0
            else:
                billsDict["AmountBill"]=sh.cell(i,ColumnAmountBill).value

            if billsDict["billQuantity"] not in filterBillsWords:
                billsTable.append(billsDict)
            i=i+1
        #print(pd.DataFrame(billsTable))
        return billsTable

    def getCoinsTable(self):
        sh=self.sh
        nameTable="Monedas"
        filterCoinsWords=["Monedas",None]
        typeCurrency=self.currency
        typeDistribution=self.distributionType
        columnsTableDict=self.indexColumns[typeDistribution]["Detalle Operaciones"][typeCurrency]["efectivo"][nameTable]
        ColumCurrency=columnsTableDict["valor"]
        ColumnCurrencyQuantity=columnsTableDict["Cantidad"]
        ColumnCurrencyAmount=columnsTableDict["subtotal"]
        ColumnTotalCurrency=columnsTableDict["total"]
        i=self.get_left_up_vertex_table(nameTable,"efectivo")
        downLimitWord=self.kwordsRowLimits[typeDistribution][typeCurrency][nameTable]["inferior"]
        coinsTable=[]
        while sh.cell(i,ColumnTotalCurrency).value !=downLimitWord:
            coinsDict={
                "coinValue":sh.cell(i,ColumCurrency).value,
                "coinQuantity":sh.cell(i,ColumnCurrencyQuantity).value
            }
            if sh.cell(i,ColumnCurrencyAmount).value=="-":
                coinsDict["AmountCoin"]=0
            else:
                coinsDict["AmountCoin"]=sh.cell(i,ColumnCurrencyAmount).value

            if coinsDict["coinValue"] not in filterCoinsWords:
                coinsTable.append(coinsDict)
            i=i+1

        #print(pd.DataFrame(coinsTable))
        return coinsTable
    def getChecksTable(self):
        sh=self.sh
        nameTable="Cheques"
        moneyType="bancario"
        filterCheckWords=["Fecha",None,"Cheques"]
        typeCurrency=self.currency
        typeDistribution=self.distributionType
        columnsTableDict=self.indexColumns[typeDistribution]["Detalle Operaciones"][typeCurrency][moneyType][nameTable]
        ColumDate=columnsTableDict["fecha"]
        ColumCheckDocument=columnsTableDict["NroDocumento"]
        ColumnCheckBank=columnsTableDict["Banco"]
        ColumCheckAmount=columnsTableDict["subtotal"]
        ColumCheckTotal=columnsTableDict["total"]
        uplimitWord=self.kwordsRowLimits[typeDistribution][typeCurrency][nameTable]["superior"]
        downLimitWord=self.kwordsRowLimits[typeDistribution][typeCurrency][nameTable]["inferior"]
        checkTable=[]
        i=1
        while sh.cell(i,ColumDate).value !=uplimitWord:
            i=i+1
        while sh.cell(i,ColumCheckTotal).value !=downLimitWord:
            checkDict={
                "DateCheck":sh.cell(i,ColumDate).value,
                "DocumentNumber":sh.cell(i,ColumCheckDocument).value,
                "CheckBank":sh.cell(i,ColumnCheckBank).value,
                "AmountCheck":""
            }
            if checkDict["AmountCheck"]=="-":
                checkDict["AmountCheck"]=0
            else:
                checkDict["AmountCheck"]=sh.cell(i,ColumCheckAmount).value
            if checkDict["DateCheck"] not in filterCheckWords:
                checkTable.append(checkDict)
            i=i+1
        #print(pd.DataFrame(checkTable))
        return checkTable
    def getBankTransfersTable(self):
        sh=self.sh
        nameTable="transferencias"
        typeCurrency=self.currency
        moneyType="bancario"
        filterTransferWords=["Transferencias y/o Depósitos","Total Depósitos","Fecha",None]
        columnsTableDict=self.indexColumns[self.distributionType]["Detalle Operaciones"][typeCurrency][moneyType][nameTable]
        ColumBankTransfer=columnsTableDict["fecha"]
        ColumBankTransferDocument=columnsTableDict["NroDocumento"]
        ColumBankTransferBank=columnsTableDict["Banco"]
        ColumBankTransferAmount=columnsTableDict["subtotal"]
        ColumBankTransferTotal=columnsTableDict["total"]
        i=1
        upLimitWord=self.kwordsRowLimits[self.distributionType][typeCurrency][nameTable]["superior"]
        downLimitWord=self.kwordsRowLimits[self.distributionType][typeCurrency][nameTable]["inferior"]
        while sh.cell(i,ColumBankTransfer).value !=upLimitWord:
            i=i+1

        bankTransferTable=[]
        while sh.cell(i,ColumBankTransferTotal).value !=downLimitWord:
            bankTransferDict={
                "DateTransfer":sh.cell(i,ColumBankTransfer).value,
                "DocumentNumberTransfer":sh.cell(i,ColumBankTransferDocument).value,
                "BankTransfer":sh.cell(i,ColumBankTransferBank).value,
                "AmountTransfer":sh.cell(i,ColumBankTransferAmount).value
            }

            if bankTransferDict["AmountTransfer"]=="-":
                bankTransferDict["AmountTransfer"]=0
            else:
                bankTransferDict["AmountTransfer"]=sh.cell(i,ColumBankTransferAmount).value
            if bankTransferDict["DateTransfer"] not in filterTransferWords:
                bankTransferTable.append(bankTransferDict)
            i=i+1
        #print(pd.DataFrame(bankTransferTable))
        return bankTransferTable
    def getSummaryTable(self):
        sh=self.sh
        columnsTableDict=self.indexColumns["distribuidora"]["Reporte Cobrador"]["ambos"]["reporte"]["reporte"]
        ColumCode=columnsTableDict["codigo"]
        ColumChecker=columnsTableDict["cobrador"]
        ColumRendDate=columnsTableDict["Fecha de Rend."]
        ColumReceiptDate=columnsTableDict["Fecha Recibo Emitido"]
        ColumReceiptNumber=columnsTableDict["NroRecibo"]
        ColumUsCash=columnsTableDict["EfectivoDolar"]
        ColumBsCash=columnsTableDict["EfectivoBs"]
        ColumUsCheck=columnsTableDict["ChequesDolar"]
        ColumBsCheck=columnsTableDict["ChequesBs"]
        ColumTotalUsCash=columnsTableDict["TotalEfectivoUs"]
        ColumTotalEqBsCash=columnsTableDict["TotalEfectivoEqBs"]
        ColumTotalBsCash=columnsTableDict["TotalEfectivoBs"]
        ColumTransferUs=columnsTableDict["TransfUs"]
        ColumTransferEqBs=columnsTableDict["TransfEqBs"]
        ColumTransferBs=columnsTableDict["TransfBs"]
        ColumTotalUs=columnsTableDict["CobradoUs"]
        ColumTotalEqBs=columnsTableDict["CobradoEqBs"]
        ColumTotalBs=columnsTableDict["CobradoBs"]
        ColumTotal=columnsTableDict["CobradoTotal"]
        
        upLimitWord=self.kwordsRowLimits["distribuidora"]["ambos"]["reporte"]["superior"]
        downLimitWord=self.kwordsRowLimits["distribuidora"]["ambos"]["reporte"]["inferior"]
        i=1
        summaryTable=[]
        filterkeyWords=[None,"Fecha de Rend."]
        while sh.cell(i,ColumRendDate).value !=upLimitWord:
            i=i+1

        while sh.cell(i,ColumRendDate).value !=downLimitWord:
            formatedDate=str(sh.cell(i,ColumReceiptDate).value).replace("/","")
            totalFormat=str(sh.cell(i,ColumTotal).value).replace(",","")
            checker=str(sh.cell(i,ColumChecker).value)
            summaryDict={
                "Code_CcajConsol":sh.cell(i,ColumCode).value,
                "Checker_CcajConsol":sh.cell(i,ColumChecker).value,
                "FechaRend_CcajConsol":sh.cell(i,ColumRendDate).value,
                "FechaRecibo_CcajConsol":sh.cell(i,ColumReceiptDate).value,
                "NroRecibo_CcajConsol":sh.cell(i,ColumReceiptNumber).value,
                "UsCash_CcajConsol":sh.cell(i,ColumUsCash).value,
                "BsCash_CcajConsol":sh.cell(i,ColumBsCash).value,
                "UsCheck_CcajConsol":sh.cell(i,ColumUsCheck).value,
                "BsCheck_CcajConsol":sh.cell(i,ColumBsCheck).value,
                "TotalUsCash_CcajConsol":sh.cell(i,ColumTotalUsCash).value,
                "TotalEqBsCash_CcajConsol":sh.cell(i,ColumTotalEqBsCash).value,
                "TotalBsCash_CcajConsol":sh.cell(i,ColumTotalBsCash).value,
                "TransferUs_CcajConsol":sh.cell(i,ColumTransferUs).value,
                "TransferEqBs_CcajConsol":sh.cell(i,ColumTransferEqBs).value,
                "TransferBs_CcajConsol":sh.cell(i,ColumTransferBs).value,
                "TotalBs_CcajConsol":sh.cell(i,ColumTotalBs).value,
                "TotalUs_CcajConsol":sh.cell(i,ColumTotalUs).value,
                "TotalEqBs_CcajConsol":sh.cell(i,ColumTotalEqBs).value,
                "TotalCcajConsol":sh.cell(i,ColumTotalBs).value,
            }
            if summaryDict["FechaRend_CcajConsol"] not in filterkeyWords:
                totalFormat="{:.2f}".format(float(totalFormat))
                rec=self.recaud
                #summaryDict["uniqKey"]=rec+"_"+checker+"_"+formatedDate+"_"+totalFormat
                if float(totalFormat)>=0:
                    summaryTable.append(summaryDict)
            i=i+1
        #print(pd.DataFrame(summaryTable))        
        return summaryTable

    def get_vouchers_table(self):
        sh=self.sh
        nameTable="voucher"
        typeCurrency=self.currency
        typeDistribution=self.distributionType
        columnstableDict=self.indexColumns["agencia"]["Detalle Operaciones"][typeCurrency]["eqEfectivo"][nameTable]
        columDate=columnstableDict["fecha"]
        columnNroRef=columnstableDict["Nro. Ref"]
        ColumNroClient=columnstableDict["Nro. CI."]
        ColumSubtotal=columnstableDict["subtotal"]
        ColumTotal=columnstableDict["total"]
        i=1
        voucherTable=[]
        kwordsFilter=[None,"Fecha","Total vouchers","Voucher de Tarjetas"]
        upLimitWord=self.kwordsRowLimits[typeDistribution][typeCurrency][nameTable]["superior"]
        downLimitWord=self.kwordsRowLimits[typeDistribution][typeCurrency][nameTable]["inferior"]
        while sh.cell(i,columDate).value !=upLimitWord:
            i=i+1
        while sh.cell(i,ColumTotal).value !=downLimitWord:
            voucherDict={
                "DateVoucher":sh.cell(i,columDate).value,
                "NroRefVoucher":sh.cell(i,columnNroRef).value,
                "NroClientVoucher":sh.cell(i,ColumNroClient).value,
                "AmountVoucher":sh.cell(i,ColumSubtotal).value,
            }
            if voucherDict["AmountVoucher"]=="-":
                voucherDict["AmountVoucher"]=0
            else:
                voucherDict["AmountVoucher"]=sh.cell(i,ColumSubtotal).value

            if voucherDict["DateVoucher"] not in kwordsFilter:
                voucherTable.append(voucherDict)
            i=i+1
        #print(pd.DataFrame(voucherTable))
        return voucherTable
    def get_coupon_table(self):
        sh=self.sh
        nameTable="vales"
        typeCurrency=self.currency
        typeDistribution=self.distributionType
        columnstableDict=self.indexColumns[typeDistribution]["Detalle Operaciones"][typeCurrency]["eqEfectivo"][nameTable]
        columQuantity=columnstableDict["Cantidad"]
        columClient=columnstableDict["Cliente"]
        ColumSubtotal=columnstableDict["subtotal"]
        ColumTotal=columnstableDict["total"]

        i=1
        couponTable=[]
        keywordsFilter=[None,"Total vales"]
        upLimitWord=self.kwordsRowLimits[typeDistribution][typeCurrency][nameTable]["superior"]
        downLimitWord=self.kwordsRowLimits[typeDistribution][typeCurrency][nameTable]["inferior"]

        while sh.cell(i,columQuantity).value !=upLimitWord:
            i=i+1
        while sh.cell(i,ColumTotal).value !=downLimitWord:
            couponDict={
                "QuantityVale":sh.cell(i,columQuantity).value,
                "ClientVale":sh.cell(i,columClient).value,
                "SubtotalVale":sh.cell(i,ColumSubtotal).value,
            }
            i=i+1
        if couponDict["SubtotalVale"]=="-":
            couponDict["SubtotalVale"]=0
        else:
            couponDict["SubtotalVale"]=sh.cell(i,ColumSubtotal).value
        if couponDict["QuantityVale"] not in keywordsFilter:
            couponTable.append(couponDict)
            
        #print(pd.DataFrame(couponTable))
        return couponTable
    def get_qr_table(self):
        sh=self.sh
        nameTable="QR"
        typeDistribution=self.distributionType
        typeCurrency=self.currency
        columnstableDict=self.indexColumns["agencia"]["Detalle Operaciones"][typeCurrency]["eqEfectivo"][nameTable]
        columDate=columnstableDict["fecha"]
        columNroRef=columnstableDict["Nro. Ref"]
        ColumNroClient=columnstableDict["Nombre"]
        ColumSubtotal=columnstableDict["subtotal"]
        ColumTotal=columnstableDict["total"]

        i=1
        qrTable=[]
        keywordsFilter=[None,"Pagos QR","Fecha"]
        if typeCurrency=="Bs":
            pass
        upLimitWord=self.kwordsRowLimits[typeDistribution][typeCurrency][nameTable]["superior"]
        downLimitWord=self.kwordsRowLimits[typeDistribution][typeCurrency][nameTable]["inferior"]
        #print(self.fileName)
        #print(upLimitWord,columDate)
        while sh.cell(i,columDate).value !=upLimitWord:
            i=i+1
        
        while sh.cell(i,ColumTotal).value !=downLimitWord:
            qrDict={
                "DateQr":sh.cell(i,columDate).value,
                "NroRefQr":sh.cell(i,columNroRef).value,
                "NroClientQr":sh.cell(i,ColumNroClient).value,
                "SubtotalQr":sh.cell(i,ColumSubtotal).value,
            }
            if qrDict["SubtotalQr"]=="-":
                qrDict["SubtotalQr"]=0
            else:
                qrDict["SubtotalQr"]=sh.cell(i,ColumSubtotal).value
            if qrDict["DateQr"] not in keywordsFilter:
                qrTable.append(qrDict)
            i=i+1
        #print(pd.DataFrame(qrTable))
        return qrTable
    def get_diferences_table(self):
        typeCurrency=self.currency
        typeDistribution=self.distributionType
        sh=self.sh
        nameTable="diferencias"
        if typeCurrency=="Bs00":
            #print(pd.DataFrame([]))
            return []
        columnstableDict=self.indexColumns[typeDistribution]["Detalle Operaciones"][typeCurrency]["contabilidad"][nameTable]
        columConcept=columnstableDict["Concepto"]
        columMotive=columnstableDict["Motivo"]
        ColumSubtotalUs=columnstableDict["subtotalDolar"]
        ColumSubtotalBs=columnstableDict["subtotalBs"]
        ColumTotalBs=columnstableDict["TotalBs"]
        ColumTotal=columnstableDict["TotalDiferencia"]

        upLimitWord=self.kwordsRowLimits[typeDistribution][typeCurrency][nameTable]["superior"]
        downLimitWord=self.kwordsRowLimits[typeDistribution][typeCurrency][nameTable]["inferior"]
        
        i=1
        diferencesTable=[]
        keywordsFilter=[None,"Total Diferencias","Concepto"]
        while sh.cell(i,columConcept).value !=upLimitWord:
            i=i+1
        while sh.cell(i,ColumTotal).value !=downLimitWord:
            diferencesDict={
                "Concept":sh.cell(i,columConcept).value,
                "Motive":sh.cell(i,columMotive).value,
                "SubtotalUs":sh.cell(i,ColumSubtotalUs).value,
                "SubtotalBs":sh.cell(i,ColumSubtotalBs).value,
                "TotalBs":sh.cell(i,ColumTotalBs).value,
            }
            
            if diferencesDict["Concept"] not in keywordsFilter:
                diferencesTable.append(diferencesDict)
            i=i+1
        #print(pd.DataFrame(diferencesTable))
        return diferencesTable
    def ventasEfectuadas(self):
        sh=self.sh
        ventasEfectuadas={}
        ventasEfectuadas["Ventas al Contado"]={}
        ventasEfectuadas["Ventas al Credito"]={}
        i=13
        j=1
        while sh.cell(i,j).value !='A':
            j=j+1
        ventasEfectuadas["Ventas al Contado"]['EFECTIVO']=sh.cell(i+1,j).value
        
        while sh.cell(i,j).value !='B':
            j=j+1
        ventasEfectuadas["Ventas al Contado"]['DOCUMENTOS']=sh.cell(i+1,j).value
        
        while sh.cell(i,j).value !='C=A+B':
            j=j+1
        ventasEfectuadas["Ventas al Contado"]['TOTAL']=sh.cell(i+1,j).value
        
        while sh.cell(i,j).value !='D':
            j=j+1   
        ventasEfectuadas["Ventas al Credito"]['PERSONAL IVSA']=''

        while sh.cell(i,j).value !='E':
            j=j+1
        ventasEfectuadas["Ventas al Credito"]['OTROS']=sh.cell(i+1,j).value
        
        while sh.cell(i,j).value !='F=D+E':
            j=j+1
        ventasEfectuadas["Ventas al Credito"]['TOTAL']=sh.cell(i+1,j).value
        
        while sh.cell(i,j).value !='G=C+F':
            j=j+1
        ventasEfectuadas['TOTALVENTAS']=sh.cell(i+1,j).value

        return ventasEfectuadas
    def otrosIngresos(self):
        sh=self.sh
        Nofounded=True
        counter=0
        i=21
        j=16
        b1=False
        b2=False
        b3=False
        while i<=sh.max_row and Nofounded:
            j=13
            while j<=20:
                #print(i,j)
                counter+=1
                if sh.cell(row=i,column=j).value=="Total Recuento Moneda Extranjera en Bs.:":
                    b1=True
                    totalMEBs=sh.cell(row=i,column=j+12).value
                if sh.cell(row=i,column=j).value=="Total Efectivo en Bs.(J=H+I):":
                    b2=True
                    totalEfectivoBs=sh.cell(row=i,column=j+12).value
                if b1 and b2:
                    Nofounded=False
                    break
                j+=1
            i+=1
        i=16
        b1=False
        Nofounded=True
        b2=False
        while i<=sh.max_row and Nofounded:
            j=16
            while j<=sh.max_column:
                #print(i,j)
                counter+=1
                if sh.cell(row=i,column=j).value=="M":
                    fondoCambios=sh.cell(row=i+1,column=j).value
                    b1=True
                if sh.cell(row=i,column=j).value=="N=J-M":
                    b2=True
                    importeDepositarBs=sh.cell(row=i+1,column=j).value
                    importeDepositarUs=sh.cell(row=i+1,column=j+4).value
                if b1 and b2:
                    Nofounded=False
                    break
                j+=1
            i+=1
            
        otrosIngresos={
            "totalMEBs":totalMEBs,
            "totalEfectivoBs":totalEfectivoBs,
            "fondosCambios":fondoCambios,
            "importeDepositarBs":importeDepositarBs,
            "importeDepositarUs":importeDepositarUs
        }
        #print(otrosIngresos)
        return otrosIngresos    
def scrapXlsxFile(fileName):
    dicts=[]
    billTable=None
    checkTable=None
    bankTransferTable=None
    voucherTable=None
    cuoponTable=None
    qrTable=None
    coinsTable=None
    diferencesTable=None
    summaryTable=None
    if fileName.find("dist")!=-1:
        distributionType="distribuidora"
    elif fileName.find("ag")!=-1:
        distributionType="agencia"
    scrapyxlsx=scrapTablesExcel(fileName,distributionType)
    fileName=Path(fileName).name
    if distributionType=="distribuidora":
        if fileName.find("first")!=-1:
            summaryTable=scrapyxlsx.getSummaryTable()
            dictsTable={"summaryTable":summaryTable}    
        elif fileName.find("Us")!=-1:
            billTable=scrapyxlsx.getBillstable()
            checkTable=scrapyxlsx.getChecksTable()
            bankTransferTable=scrapyxlsx.getBankTransfersTable()
            dictsTable={"billTable":billTable,"checkTable":checkTable,"bankTransferTable":bankTransferTable}
        elif fileName.find("Bs")!=-1:
            billTable=scrapyxlsx.getBillstable()
            coinsTable=scrapyxlsx.getCoinsTable()
            checkTable=scrapyxlsx.getChecksTable()
            bankTransferTable=scrapyxlsx.getBankTransfersTable()
            dictsTable={"billTable":billTable,"coinsTable":coinsTable,"checkTable":checkTable,"bankTransferTable":bankTransferTable}
        else:
            print("Error en el nombre del archivo")
    elif distributionType=="agencia":
        if fileName.find("Us")!=-1:
            billTable=scrapyxlsx.getBillstable()
            voucherTable=scrapyxlsx.get_vouchers_table()
            qrTable=scrapyxlsx.get_qr_table()
            dictsTable={"billTable":billTable,"voucherTable":voucherTable,"qrTable":qrTable}
        elif fileName.find("Bs")!=-1:
            scrapyxlsx.updateCurrency()
            billTable=scrapyxlsx.getBillstable()
            coinsTable=scrapyxlsx.getCoinsTable()
            voucherTable=scrapyxlsx.get_vouchers_table()
            cuoponTable=scrapyxlsx.get_coupon_table()
            qrTable=scrapyxlsx.get_qr_table()
            diferencesTable=scrapyxlsx.get_diferences_table()
            ventas=scrapyxlsx.ventasEfectuadas()
            otrosingresos=scrapyxlsx.otrosIngresos()
            dictsTable={"billTable":billTable,"coinsTable":coinsTable,"voucherTable":voucherTable,"cuoponTable":cuoponTable,"qrTable":qrTable,"diferencesTable":diferencesTable,"ventas":ventas,"otrosIngresos":otrosingresos}
        else:
            print("Error en el nombre del archivo")


    if billTable!=None:
        billst.extend(billTable)
    if checkTable!=None:
        checkstable.extend(checkTable)
    if bankTransferTable!=None:
        bankTransferstable.extend(bankTransferTable)
    if coinsTable!=None:
        coinssTable.extend(coinsTable)
    if voucherTable!=None:
        vouchersTable.extend(voucherTable)
    if qrTable!=None:
        qrsTable.extend(qrTable)
    if cuoponTable!=None:
        cuoponsTable.extend(cuoponTable)
    if diferencesTable!=None:
        diferencessTable.extend(diferencesTable)
    if summaryTable!=None:
        summariesTable.extend(summaryTable)
    dataReturn={"distributionType":distributionType,"data":dictsTable,"typeMoney":scrapyxlsx.currency}
    return dataReturn

def scrapCierresDeCaja2():
    print("Procesando archivos de cierres de caja")
    #list of .xlsx files in the directory
    xlsxFilesList=[x for x in os.listdir(r"Cierres de Caja\formatoxlsx") if x.endswith(".xlsx")]
    global billst,checkstable,bankTransferstable,coinssTable,vouchersTable,qrsTable,cuoponsTable,diferencessTable,summariesTable
    billst=[]
    checkstable=[]
    bankTransferstable=[]
    vouchersTable=[]
    coinssTable=[]
    vouchersTable=[]
    qrsTable=[]
    cuoponsTable=[]
    diferencessTable=[]
    summariesTable=[]
    for xlsxFile in xlsxFilesList:
        print(xlsxFile)
        vd=scrapXlsxFile(xlsxFile)
    if len(xlsxFilesList)==0:
        print("No hay archivos xlsx en la carpeta Cierres de Caja")
        return
    df_bt=pd.DataFrame(billst)
    df_bt.to_csv(os.path.join(get_tables_path(),"billsTable.csv"),index=False,sep=";")

    df_checks=pd.DataFrame(checkstable)
    df_checks.to_csv(os.path.join(get_tables_path(),"checksTable.csv"),index=False,sep=";")

    df_bankTransfers=pd.DataFrame(bankTransferstable)
    df_bankTransfers.to_csv(os.path.join(get_tables_path(),"bankTransfersTable.csv"),index=False,sep=";")

    df_coins=pd.DataFrame(coinssTable)
    df_coins.to_csv(os.path.join(get_tables_path(),"coinsTable.csv"),index=False,sep=";")

    df_voucher=pd.DataFrame(vouchersTable)
    df_voucher.to_csv(os.path.join(get_tables_path(),"voucherTable.csv"),index=False,sep=";")

    df_qr=pd.DataFrame(qrsTable)
    df_qr.to_csv(os.path.join(get_tables_path(),"qrTable.csv"),index=False,sep=";")

    df_cuopon=pd.DataFrame(cuoponsTable)
    df_cuopon.to_csv(os.path.join(get_tables_path(),"cuoponTable.csv"),index=False,sep=";")

    df_summaries=pd.DataFrame(summariesTable)
    df_summaries.to_csv(os.path.join(get_tables_path(),"summariesTable.csv"),index=False,sep=";")

    normalizeTable()
    #print("SCRAP CIERRES DE CAJA TERMINADO")

def scrapCierresDeCaja():
    print("-------------Procesando archivos de cierres de caja...")
    pathCashClosingInfo=os.path.join(paths.folderProyect,r"src\target\CashClosingInfo.json")
    #print(pathCashClosingInfo)
    with open(pathCashClosingInfo,"r") as json_file:
        data = json.load(json_file)
    global billst,checkstable,bankTransferstable,coinssTable,vouchersTable,qrsTable,cuoponsTable,diferencessTable,summariesTable
    
    billst=[]
    checkstable=[]
    bankTransferstable=[]
    vouchersTable=[]
    coinssTable=[]
    vouchersTable=[]
    qrsTable=[]
    cuoponsTable=[]
    diferencessTable=[]
    summariesTable=[]
    
    for i,row in enumerate(data['data']):
        for j,path in enumerate(row["xlsFilesList"]):
            if path['descargado']=="OK":
                try:
                    path=os.path.join(paths.dirCcaj,path['name']+".xlsx")
                    vd=scrapXlsxFile(path)
                    #data['data'][i]["xlsFilesList"][j]={}
                    data['data'][i]["xlsFilesList"][j]["file"]=path
                    data['data'][i]["xlsFilesList"][j]["distributionType"]=vd["distributionType"]
                    data['data'][i]["xlsFilesList"][j]["moneyType"]=vd["typeMoney"]
                    data['data'][i]["xlsFilesList"][j]["data"]=vd["data"]
                except Exception as e:
                    print(e)
                #print(row['Código'])
            #except:
                #pass

    df_bt=pd.DataFrame(billst)
    df_bt.to_csv(os.path.join(get_tables_path(),"billsTable.csv"),index=False,sep=";")

    df_checks=pd.DataFrame(checkstable)
    df_checks.to_csv(os.path.join(get_tables_path(),"checksTable.csv"),index=False,sep=";")

    df_bankTransfers=pd.DataFrame(bankTransferstable)
    df_bankTransfers.to_csv(os.path.join(get_tables_path(),"bankTransfersTable.csv"),index=False,sep=";")

    df_coins=pd.DataFrame(coinssTable)
    df_coins.to_csv(os.path.join(get_tables_path(),"coinsTable.csv"),index=False,sep=";")

    df_voucher=pd.DataFrame(vouchersTable)
    df_voucher.to_csv(os.path.join(get_tables_path(),"voucherTable.csv"),index=False,sep=";")

    df_qr=pd.DataFrame(qrsTable)
    df_qr.to_csv(os.path.join(get_tables_path(),"qrTable.csv"),index=False,sep=";")

    df_cuopon=pd.DataFrame(cuoponsTable)
    df_cuopon.to_csv(os.path.join(get_tables_path(),"cuoponTable.csv"),index=False,sep=";")

    df_summaries=pd.DataFrame(summariesTable)
    df_summaries.to_csv(os.path.join(get_tables_path(),"summariesTable.csv"),index=False,sep=";")

    normalizeTable()
    print("-------------Proceso de archivos de cierres de caja terminado")
    with open(paths.FullExcelDataJson, 'w') as outfile:
        json.dump(data, outfile,indent=4)
if __name__ == "__main__":
    scrapCierresDeCaja()
    #scrapCierresDeCaja()
    #scrapCierresDeCobrador()
    #test_scrapDolarOperationsXls()
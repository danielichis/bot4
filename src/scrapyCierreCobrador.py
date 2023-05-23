import openpyxl
from utils import get_current_path,get_index_columns_config,get_currency,get_kwords_rowLimits_config,configToJson,get_tables_path
from utils import convert_xls,pathsProyect
import os
import re
import json
import pandas as pd

pths=pathsProyect()
class scraperCierreCobrador():
    def __init__(self,fileName) -> None:
        self.XlsxPath=os.path.join(get_current_path(),"Cierres de Cobrador","formatoxlsx",fileName)
        self.fileName=fileName
        self.indexColumns=get_index_columns_config()
        self.kwordsRowLimits=get_kwords_rowLimits_config()
        self.sh=openpyxl.load_workbook(self.XlsxPath).worksheets[0]
        self.recaud=None
        self.lastRow=1
        self.getRecaud()
    
    def getRecaud(self):
        self.recaud=re.findall(r'(.*?)_', self.fileName)[0]
    def ClientToCollectorTable(self):

        tableColumns=self.indexColumns['distribuidora']['Cierre de cobrador']
        tableKwords=self.kwordsRowLimits['distribuidora']['ambos']['recibo de cobranza']
        upperLimit=tableKwords['superior']
        lowerLimit=tableKwords['inferior']
        leftColumn=tableColumns['ambos']['recibo de caja']['recibo de caja']['Nro APP']

        appNumber=leftColumn
        reciboDeCajaTable=[]
        filtersKwords=['Nro APP','Nº de APP','Datos del Recibo de Cobranza']
        i=8
        j=1
        
        while j<self.sh.max_column:
            if self.sh.cell(row=i,column=j).value=="Datos del Recibo de Cobranza":
                appNumber=j
                j+=1
                while self.sh.cell(row=i,column=j).value==None:
                    if self.sh.cell(row=i+1,column=j).value=='Fecha':
                        recepitDate=j
                    j+=1
            if self.sh.cell(row=i,column=j).value=='Datos del Cliente':
                clientCode=j
                j+=1
                while self.sh.cell(row=i,column=j).value==None:
                    if self.sh.cell(row=i+1,column=j).value=='Nombre':
                        clientName=j
                    j+=1
            if self.sh.cell(row=i,column=j).value=='Total Cobrado':
                bsAmount=j
                j+=1
                while self.sh.cell(row=i+1,column=j).value==None:

                    if self.sh.cell(row=i+2,column=j).value=='U$':
                        UsAmount=j
                    j+=1
            if self.sh.cell(row=i+1,column=j).value=='Cheques':
                checkDate=j
                j+=1
                while self.sh.cell(row=i+1,column=j).value==None:
                    if self.sh.cell(row=i+2,column=j).value=='Nº':
                        checkNumber=j
                    if self.sh.cell(row=i+2,column=j).value=='Banco':
                        checkBank=j
                    if self.sh.cell(row=i+2,column=j).value=='Bs.':
                        bsCheck=j
                    if self.sh.cell(row=i+2,column=j).value=='U$':
                        usCheck=j
                    j+=1
            if self.sh.cell(row=i+1,column=j).value=='Transferencia y/o Depósito':
                dateTransfer=j
                j+=1
                while self.sh.cell(row=i+1,column=j).value==None:
                    if self.sh.cell(row=i+2,column=j).value=='Banco':
                        bankTransfer=j
                    if self.sh.cell(row=i+2,column=j).value=='Bs.':
                        bsTransfer=j
                    if self.sh.cell(row=i+2,column=j).value=='U$':
                        usTransfer=j
                    j+=1
            if self.sh.cell(row=i+1,column=j).value=='Total Recaudado':
                subtotalBs=j
                j+=1
                while self.sh.cell(row=i+1,column=j).value==None:
                    if self.sh.cell(row=i+2,column=j).value=='U$':
                        subtotalUs=j
                    if self.sh.cell(row=i+2,column=j).value=='Eq. Bs.':
                        subtotalEqBs=j
                    j+=1
            if self.sh.cell(row=i+1,column=j).value=='Total':
                totalBs=j
                j+=1
            j+=1
            
        while self.sh.cell(row=i,column=clientCode).value!=lowerLimit and self.sh.cell(row=i,column=6).value!=lowerLimit:
            ditTable={
                'ruta_CcajCobClient':self.fileName[:-5],
                "recaudadora_CcajCobClient":self.recaud,
                'Nro APP_CcajCobClient':self.sh.cell(row=i,column=appNumber).value,
                'Fecha Recibo_CcajCobClient':self.sh.cell(row=i,column=recepitDate).value,
                'Cod Cliente_CcajCobClient':self.sh.cell(row=i,column=clientCode).value,
                'Nombre cliente_CcajCobClient':self.sh.cell(row=i,column=clientName).value,
                'CashBs_CcajCobClient':self.sh.cell(row=i,column=bsAmount).value,
                'CashUs_CcajCobClient':self.sh.cell(row=i,column=UsAmount).value,
                'CheckDate_CcajCobClient':self.sh.cell(row=i,column=checkDate).value,
                'CheckNumber_CcajCobClient':self.sh.cell(row=i,column=checkNumber).value,
                'CheckBank_CcajCobClient':self.sh.cell(row=i,column=checkBank).value,
                'CheckBs_CcajCobClient':self.sh.cell(row=i,column=bsCheck).value,
                'CheckUs_CcajCobClient':self.sh.cell(row=i,column=usCheck).value,
                'TransferDate_CcajCobClient':self.sh.cell(row=i,column=dateTransfer).value,
                'TransferBank_CcajCobClient':self.sh.cell(row=i,column=bankTransfer).value,
                'TransferBs_CcajCobClient':self.sh.cell(row=i,column=bsTransfer).value,
                'TransferUs_CcajCobClient':self.sh.cell(row=i,column=usTransfer).value,
                'SubtotalBs_CcajCobClient':self.sh.cell(row=i,column=subtotalBs).value,
                'SubtotalUs_CcajCobClient':self.sh.cell(row=i,column=subtotalUs).value,
                'SubtotalEqBs_CcajCobClient':self.sh.cell(row=i,column=subtotalEqBs).value,
                'Total_CcajCobClient':0.00,
            }

            if self.sh.cell(row=i,column=appNumber).value!=None and self.sh.cell(row=i,column=appNumber).value not in filtersKwords:
                ditTable["Total_CcajCobClient"]="{:.2f}".format(float(self.sh.cell(row=i,column=totalBs).value))
                reciboDeCajaTable.append(ditTable)
            if self.sh.cell(row=i+1,column=appNumber+1).value=="Datos del Recibo de Cobranza":
                appNumber+=1
                recepitDate+=1
                clientCode+=1
                clientName+=1
                bsAmount+=1
                UsAmount+=1
                checkDate+=1
                checkNumber+=1
                checkBank+=1
                bsCheck+=1
                usCheck+=1
                dateTransfer+=1
                bankTransfer+=1
                bsTransfer+=1
                usTransfer+=1
                subtotalBs+=1
                subtotalUs+=1
                subtotalEqBs+=1
                totalBs+=1
            i+=1
        self.lastRow=i
        #print(pd.DataFrame(reciboDeCajaTable))
        return reciboDeCajaTable
    def CollectorToBoxTable(self):
        tableKwords=self.kwordsRowLimits['distribuidora']['ambos']['recepcion en caja']
        upperLimit=tableKwords['superior']
        botomLimit=tableKwords['inferior']
        i=1
        j=11
        # while j<self.sh.max_row:
        #     if self.sh.cell(row=j,column=4).value=="Total efectivo recibido":
        #         pass
        #     if self.sh.cell(row=j,column=5).value=="Total efectivo recibido":
        #         pass
        #     j+=1
        while  self.sh.cell(row=i,column=11).value!=upperLimit:
            if self.sh.cell(row=i,column=12).value==upperLimit:
                j=12
                break
            i+=1
        filtersKwords=['Recepción en caja',"Efectivo","Bs."]

        while j<self.sh.max_column:
            if self.sh.cell(row=i+1,column=j).value=="Efectivo":
                cashBs=j
                j=j+1
                while self.sh.cell(row=i+1,column=j).value==None:
                    if self.sh.cell(row=i+2,column=j).value=="U$":
                        cashUs=j
                    if self.sh.cell(row=i+2,column=j).value=="Eq. Bs.":
                        cashEqBs=j
                    j+=1
            if self.sh.cell(row=i+1,column=j).value=="Cheques":
                checkBs=j
                j=j+1
                while self.sh.cell(row=i+1,column=j).value==None:
                    if self.sh.cell(row=i+2,column=j).value=="U$":
                        checkUs=j
                    if self.sh.cell(row=i+2,column=j).value=="Eq. Bs.":
                        checkEqBs=j
                    j+=1
            if self.sh.cell(row=i+1,column=j).value=="Transf. y/o Dep.":
                dateTransfer=j
                j=j+1
                while self.sh.cell(row=i+1,column=j).value==None:
                    if self.sh.cell(row=i+2,column=j).value=="Banco":
                        bankTransfer=j
                    if self.sh.cell(row=i+2,column=j).value=="Bs.":
                        bsTransfer=j
                    if self.sh.cell(row=i+2,column=j).value=="U$":
                        usTransfer=j
                    if self.sh.cell(row=i+2,column=j).value=="Eq. Bs.":
                        EqBsTransfer=j
                    j+=1
            if self.sh.cell(row=i+2,column=j).value=="Bs.":
                totalBs=j
            j+=1
                
        receiptBoxTable=[]
        downRow=self.lastRow+1
        j=1
        while self.sh.cell(row=downRow,column=j).value!="Recepción en caja":
            j=j+1
        firstRowWithData=downRow+3
        i=firstRowWithData
        while self.sh.cell(row=i,column=j).value!="Cargos al Cobrador" and self.sh.cell(row=i,column=j+1).value!="Cargos al Cobrador":
            ditTable={
                'ruta_CcajCobCob':self.fileName[:-5],
                "recaudadora_CcajCobCob":self.recaud,
                "CashBs_CcajCobCob":self.sh.cell(row=i,column=cashBs).value,
                "CashUs_CcajCobCob":self.sh.cell(row=i,column=cashUs).value,
                "CashEqBs_CcajCobCob":self.sh.cell(row=i,column=cashEqBs).value,
                "CheckBs_CcajCobCob":self.sh.cell(row=i,column=checkBs).value,
                "CheckUs_CcajCobCob":self.sh.cell(row=i,column=checkUs).value,
                "CheckEqBs_CcajCobCob":self.sh.cell(row=i,column=checkEqBs).value,
                "TransferDate_CcajCobCob":self.sh.cell(row=i,column=dateTransfer).value,
                "TransferBank_CcajCobCob":self.sh.cell(row=i,column=bankTransfer).value,
                "TransferBs_CcajCobCob":self.sh.cell(row=i,column=bsTransfer).value,
                "TransferUs_CcajCobCob":self.sh.cell(row=i,column=usTransfer).value,
                "TransferEqBs_CcajCobCob":self.sh.cell(row=i,column=EqBsTransfer).value,
                "TotalCCOBCAJA_CcajCobCob":0.00 
            }
            if  self.sh.cell(row=i,column=cashBs).value!=None and self.sh.cell(row=i,column=cashBs).value not in filtersKwords:
                value=str(self.sh.cell(row=i,column=totalBs).value).replace(",","")
                if value!="None":
                    value=float(value)
                    ditTable['TotalCCOBCAJA']="{:.2f}".format(value)
                receiptBoxTable.append(ditTable)
            if self.sh.cell(row=i,column=cashBs+1).value=="Efectivo":
                cashBs+=1
                cashUs+=1
                cashEqBs+=1
                checkBs+=1
                checkUs+=1
                checkEqBs+=1
                dateTransfer+=1
                bankTransfer+=1
                bsTransfer+=1
                usTransfer+=1
                EqBsTransfer+=1
                totalBs+=1
            i+=1
        #print(pd.DataFrame(receiptBoxTable))
        return receiptBoxTable
def scrap_CierreCobrador():
    print("-------------Procesando cierres de cobrador...")
    cierreCobradorFiles=os.listdir(os.path.join(get_current_path(),"Cierres de Cobrador","formatoxlsx"))
    collectorClientTable=[]
    collectorBoxTable=[]
    for file in cierreCobradorFiles:
        if file.endswith(".xlsx"):
            print("Procesando archivo: ",file)
            scob=scraperCierreCobrador(file)
            q=scob.ClientToCollectorTable()
            collectorClientTable.extend(q)
            p=scob.CollectorToBoxTable()
            collectorBoxTable.extend(p)
            #print("--------------------------Archivo procesado: ")
    if len(cierreCobradorFiles)==0:
        print("No hay archivos de cierres de cobrador para procesar")
        return
    df1=pd.DataFrame(collectorClientTable)
    df2=pd.DataFrame(collectorBoxTable)
    df1.to_csv(os.path.join(get_tables_path(),"collectorClientTable.csv"),index=False,sep=";")
    #save to json file
    with open(pths.jsonClientBox, 'w') as file:
        json.dump(collectorClientTable, file,indent=4)

    df2.to_csv(os.path.join(get_tables_path(),"collectorBoxTable.csv"),index=False,sep=";")
    with open(pths.jsonCobBox, 'w') as file:
        json.dump(collectorBoxTable, file,indent=4)
    print("-------------Cierres de cobrador procesados exitosamente\n")
if __name__ == "__main__":
    #collectorClosingFolder=os.path.join(get_current_path(),"Cierres de Cobrador")
    #convert_xls(collectorClosingFolder)
    scrap_CierreCobrador()
            
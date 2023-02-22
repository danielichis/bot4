import openpyxl
from utils import get_current_path,get_index_columns_config,get_currency,get_kwords_rowLimits_config
import os
import re
import json
import pandas as pd

class scrapTablesExcel:
    def __init__(self,fileName,distributionType) -> None:
        self.XlsxPath=os.path.join(get_current_path(),"descargasXlsx")
        self.indexColumns=get_index_columns_config()
        self.kwordsRowLimits=get_kwords_rowLimits_config()
        self.sh=openpyxl.load_workbook(os.path.join(get_current_path(),"descargasXlsx",fileName)).worksheets[0]
        self.currency=get_currency(fileName)
        self.distributionType=distributionType
        self.gap=0
    def get_left_up_vertex_table(self,tableName,moneyType):
        sh=self.sh
        i=1
        columBill=self.indexColumns[self.distributionType]["Detalle Operaciones"][self.currency][moneyType][tableName]["valor"]
        kword=self.kwordsRowLimits[self.distributionType][self.currency][tableName]["superior"]
        while sh.cell(i,columBill).value !=kword:
            i=i+1
        return i
    def updateCurrency(self):
        sh=self.sh
        j=1
        i=6
        #move to the right until find a number
        while (sh.cell(i,j).value==None):
            j=j+1
        if j==2:
            self.currency="Bs00"
        elif j==4:
            self.currency="Bs"
        else:
            print("Error: no se pudo determinar el gap")
    
    def getBillstable(self):
        sh=self.sh
        filterBillsWords=["Cantidad",None]
        typeCurrency=self.currency
        typeFile=""
        typeDistribution=self.distributionType
        nameTable="billetes"
        columnsTableDict=self.indexColumns[typeDistribution]["Detalle Operaciones"][typeCurrency]["efectivo"][nameTable]
        columBill=columnsTableDict["valor"]-self.gap
        columQantityBill=columnsTableDict["Cantidad"]-self.gap
        ColumnAmountBill=columnsTableDict["subtotal"]-self.gap
        columTotalBill=columnsTableDict["total"]-self.gap
        billsTable=[]
        downLimitWord=self.kwordsRowLimits[typeDistribution][typeCurrency][nameTable]["inferior"]
        moneyType="efectivo"
        i=self.get_left_up_vertex_table(nameTable,moneyType)
        while sh.cell(i,columTotalBill).value !=downLimitWord:
            billsDict={
                "billValue":sh.cell(i,columBill).value,
                "billQuantity":sh.cell(i,columQantityBill).value,
                "Amount":sh.cell(i,ColumnAmountBill).value
            }
            if billsDict["billQuantity"] not in filterBillsWords:
                billsTable.append(billsDict)
            i=i+1
        print(pd.DataFrame(billsTable))
        return pd.DataFrame(billsTable)

    def getCoinsTable(self):
        sh=self.sh
        nameTable="Monedas"
        filterCoinsWords=["Monedas",None]
        typeCurrency=self.currency
        typeDistribution=self.distributionType
        columnsTableDict=self.indexColumns[typeDistribution]["Detalle Operaciones"][typeCurrency]["efectivo"][nameTable]
        ColumCurrency=columnsTableDict["valor"]
        ColumnCurrencyQuantity=columnsTableDict["Cantidad"]
        ColumnCurrencyAmount=columnsTableDict["subtotal"]
        ColumnTotalCurrency=columnsTableDict["total"]
        i=self.get_left_up_vertex_table(nameTable,"efectivo")
        coinsTable=[]
        while sh.cell(i,ColumnTotalCurrency).value !=self.kwordsRowLimits[typeDistribution][typeCurrency][nameTable]["inferior"]:
            coinsDict={
                "coinValue":sh.cell(i,ColumCurrency).value,
                "coinQuantity":sh.cell(i,ColumnCurrencyQuantity).value,
                "Amount":sh.cell(i,ColumnCurrencyAmount).value
            }
            if coinsDict["coinValue"] not in filterCoinsWords:
                coinsTable.append(coinsDict)
            i=i+1

        print(pd.DataFrame(coinsTable))
        return pd.DataFrame(coinsTable)
    def getChecksTable(self):
        sh=self.sh
        nameTable="Cheques"
        filterCheckWords=["Cheques",None]
        typeCurrency=self.currency
        columnsTableDict=self.indexColumns[typeCurrency][nameTable]
        ColumCheck=columnsTableDict["fecha"]
        ColumCheckDocument=columnsTableDict["Nrodocumento"]
        ColumnCheckBank=columnsTableDict["Banco"]
        ColumCheckAmount=columnsTableDict["subtotal"]
        ColumCheckTotal=columnsTableDict["total"]
        i=self.get_left_up_vertex_table(nameTable)
        checkTable=[]
        while sh.cell(i,ColumCheckTotal).value !=self.kwordsRowLimits[typeCurrency][nameTable]["inferior"]:
            checkDict={
                "Date":sh.cell(i,ColumCheck).value,
                "DocumentNumber":sh.cell(i,ColumCheckDocument).value,
                "Bank":sh.cell(i,ColumnCheckBank).value,
                "Amount":sh.cell(i,ColumCheckAmount).value
            }
            if checkDict["Date"] not in filterCheckWords:
                checkTable.append(checkDict)
            i=i+1
        print(pd.DataFrame(checkTable))
        return pd.DataFrame(checkTable)
    def getBankTransfersTable(self):
        sh=self.sh
        nameTable="transferencias"
        typeCurrency=self.currency
        filterTransferWords=["Transferencias y/o Depósitos","Total Depósitos","Fecha",None]
        columnsTableDict=self.indexColumns[typeCurrency][nameTable]
        ColumBankTransfer=columnsTableDict["fecha"]
        ColumBankTransferDocument=columnsTableDict["Nrodocumento"]
        ColumBankTransferBank=columnsTableDict["Banco"]
        ColumBankTransferAmount=columnsTableDict["subtotal"]
        ColumBankTransferTotal=columnsTableDict["total"]
        i=self.get_left_up_vertex_table(nameTable)

        bankTransferTable=[]
        while sh.cell(i,ColumBankTransferTotal).value !=self.kwordsRowLimits[typeCurrency][nameTable]["inferior"]:
            bankTransferDict={
                "Date":sh.cell(i,ColumBankTransfer).value,
                "DocumentNumber":sh.cell(i,ColumBankTransferDocument).value,
                "Bank":sh.cell(i,ColumBankTransferBank).value,
                "Amount":sh.cell(i,ColumBankTransferAmount).value
            }
            if bankTransferDict["Date"] not in filterTransferWords:
                bankTransferTable.append(bankTransferDict)
            i=i+1
        print(pd.DataFrame(bankTransferTable))
        return pd.DataFrame(bankTransferTable)
    def getSummaryTable(self):
        sh=self.sh
        columnsTableDict=self.indexColumns["Distribuidora"]["Reporte Cobrador"]["ambos"]["reporte"]["reporte"]
        ColumCode=columnsTableDict["codigo"]
        ColumChecker=columnsTableDict["cobrador"]
        ColumRendDate=columnsTableDict["Fecha de Rend."]
        ColumReceiptDate=columnsTableDict["Fecha Recibo Emitido"]
        ColumReceiptNumber=columnsTableDict["NroRecibo"]
        ColumUsCash=columnsTableDict["EfectivoDolar"]
        ColumBsCash=columnsTableDict["EfectivoBs"]
        ColumUsCheck=columnsTableDict["ChequesDolar"]
        ColumBsCheck=columnsTableDict["ChequesBs"]
        ColumTotalUsCash=columnsTableDict["TotalEfectivoUs"]
        ColumTotalEqBsCash=columnsTableDict["TotalEfectivoEqBs"]
        ColumTotalBsCash=columnsTableDict["TotalEfectivoBs"]
        ColumTransferUs=columnsTableDict["TransfUs"]
        ColumTransferEqBs=columnsTableDict["TransfEqBs"]
        ColumTransferBs=columnsTableDict["TransfBs"]
        ColumTotalUs=columnsTableDict["CobradoUs"]
        ColumTotalEqBs=columnsTableDict["CobradoEqBs"]
        ColumTotalBs=columnsTableDict["CobradoBs"]
        ColumTotal=columnsTableDict["CobradoTotal"]
        
        upLimitWord=self.kwordsRowLimits["ambos"]["superior"]
        downLimitWord=self.kwordsRowLimits["ambos"]["inferior"]
        i=1
        summaryTable=[]
        filterkeyWords=[None,"Fecha de Rend."]
        while sh.cell(i,ColumRendDate).value !=upLimitWord:
            i=i+1

        while sh.cell(i,ColumRendDate).value !=downLimitWord:
            summaryDict={
                "Code":sh.cell(i,ColumCode).value,
                "Checker":sh.cell(i,ColumChecker).value,
                "FechaRend":sh.cell(i,ColumRendDate).value,
                "FechaRecibo":sh.cell(i,ColumReceiptDate).value,
                "NroRecibo":sh.cell(i,ColumReceiptNumber).value,
                "UsCash":sh.cell(i,ColumUsCash).value,
                "BsCash":sh.cell(i,ColumBsCash).value,
                "UsCheck":sh.cell(i,ColumUsCheck).value,
                "BsCheck":sh.cell(i,ColumBsCheck).value,
                "TotalUsCash":sh.cell(i,ColumTotalUsCash).value,
                "TotalEqBsCash":sh.cell(i,ColumTotalEqBsCash).value,
                "TotalBsCash":sh.cell(i,ColumTotalBsCash).value,
                "TransferUs":sh.cell(i,ColumTransferUs).value,
                "TransferEqBs":sh.cell(i,ColumTransferEqBs).value,
                "TransferBs":sh.cell(i,ColumTransferBs).value,
                "TotalUs":sh.cell(i,ColumTotalUs).value,
                "TotalEqBs":sh.cell(i,ColumTotalEqBs).value,
                "TotalBs":sh.cell(i,ColumTotalBs).value,
                "Total":sh.cell(i,ColumTotal).value
            }
            if summaryDict["FechaRend"] not in filterkeyWords:
                summaryTable.append(summaryDict)
            i=i+1        
        return pd.DataFrame(summaryTable)

    def get_vouchers_table(self):
        sh=self.sh
        nameTable="voucher"
        typeCurrency=self.currency
        typeDistribution=self.distributionType
        columnstableDict=self.indexColumns["agencia"]["Detalle Operaciones"][typeCurrency]["eqEfectivo"][nameTable]
        columDate=columnstableDict["fecha"]-self.gap
        columnNroRef=columnstableDict["Nro. Ref"]-self.gap
        ColumNroClient=columnstableDict["Nro. CI."]-self.gap
        ColumSubtotal=columnstableDict["subtotal"]-self.gap
        ColumTotal=columnstableDict["total"]-self.gap
        i=1
        voucherTable=[]
        kwordsFilter=[None,"Fecha","Total vouchers","Voucher de Tarjetas"]
        upLimitWord=self.kwordsRowLimits[typeDistribution][typeCurrency][nameTable]["superior"]
        downLimitWord=self.kwordsRowLimits[typeDistribution][typeCurrency][nameTable]["inferior"]
        while sh.cell(i,columDate).value !=upLimitWord:
            i=i+1
        while sh.cell(i,ColumTotal).value !=downLimitWord:
            voucherDict={
                "Date":sh.cell(i,columDate).value,
                "NroRef":sh.cell(i,columnNroRef).value,
                "NroClient":sh.cell(i,ColumNroClient).value,
                "Subtotal":sh.cell(i,ColumSubtotal).value,
                "Total":sh.cell(i,ColumTotal).value
            }
            if voucherDict["Date"] not in kwordsFilter:
                voucherTable.append(voucherDict)
            i=i+1
        print(pd.DataFrame(voucherTable))
        return pd.DataFrame(voucherTable)
    def get_coupon_table(self):
        sh=self.sh
        nameTable="vales"
        typeCurrency=self.currency
        typeDistribution=self.distributionType
        columnstableDict=self.indexColumns[typeDistribution]["Detalle Operaciones"][typeCurrency]["eqEfectivo"][nameTable]
        columQuantity=columnstableDict["Cantidad"]
        columClient=columnstableDict["Cliente"]
        ColumSubtotal=columnstableDict["subtotal"]
        ColumTotal=columnstableDict["total"]

        i=1
        couponTable=[]
        keywordsFilter=[None,"Total vales"]
        upLimitWord=self.kwordsRowLimits[typeDistribution][typeCurrency][nameTable]["superior"]
        downLimitWord=self.kwordsRowLimits[typeDistribution][typeCurrency][nameTable]["inferior"]

        while sh.cell(i,columQuantity).value !=upLimitWord:
            i=i+1
        while sh.cell(i,ColumTotal).value !=downLimitWord:
            couponDict={
                "Quantity":sh.cell(i,columQuantity).value,
                "Client":sh.cell(i,columClient).value,
                "Subtotal":sh.cell(i,ColumSubtotal).value,
            }
            i=i+1
        if couponDict["Quantity"] not in keywordsFilter:
            couponTable.append(couponDict)
            
        print(pd.DataFrame(couponTable))
        return pd.DataFrame(couponTable)
    def get_qr_table(self):
        sh=self.sh
        nameTable="QR"
        typeDistribution=self.distributionType
        typeCurrency=self.currency
        columnstableDict=self.indexColumns["agencia"]["Detalle Operaciones"][typeCurrency]["eqEfectivo"][nameTable]
        columDate=columnstableDict["fecha"]
        columNroRef=columnstableDict["Nro. Ref"]
        ColumNroClient=columnstableDict["Nombre"]
        ColumSubtotal=columnstableDict["subtotal"]
        ColumTotal=columnstableDict["total"]

        i=1
        qrTable=[]
        keywordsFilter=[None,"Pagos QR","Fecha"]
        upLimitWord=self.kwordsRowLimits[typeDistribution][typeCurrency][nameTable]["superior"]
        downLimitWord=self.kwordsRowLimits[typeDistribution][typeCurrency][nameTable]["inferior"]
        while sh.cell(i,columDate).value !=upLimitWord:
            i=i+1
        
        while sh.cell(i,ColumTotal).value !=downLimitWord:
            qrDict={
                "Date":sh.cell(i,columDate).value,
                "NroRef":sh.cell(i,columNroRef).value,
                "NroClient":sh.cell(i,ColumNroClient).value,
                "Subtotal":sh.cell(i,ColumSubtotal).value,
            }
            if qrDict["Date"] not in keywordsFilter:
                qrTable.append(qrDict)
            i=i+1
        print(pd.DataFrame(qrTable))
        return pd.DataFrame(qrTable)
    def get_diferences_table(self):
        typeCurrency=self.currency
        typeDistribution=self.distributionType
        sh=self.sh
        nameTable="diferencias"
        if typeCurrency=="Bs00":
            print(pd.DataFrame([]))
            return pd.DataFrame([])
        columnstableDict=self.indexColumns[typeDistribution]["Detalle Operaciones"][typeCurrency]["contabilidad"][nameTable]
        columConcept=columnstableDict["Concepto"]
        columMotive=columnstableDict["Motivo"]
        ColumSubtotalUs=columnstableDict["subtotalDolar"]
        ColumSubtotalBs=columnstableDict["subtotalBs"]
        ColumTotalBs=columnstableDict["TotalBs"]
        ColumTotal=columnstableDict["TotalDiferencia"]

        upLimitWord=self.kwordsRowLimits[typeDistribution][typeCurrency][nameTable]["superior"]
        downLimitWord=self.kwordsRowLimits[typeDistribution][typeCurrency][nameTable]["inferior"]
        
        i=1
        diferencesTable=[]
        keywordsFilter=[None,"Total Diferencias","Concepto"]
        while sh.cell(i,columConcept).value !=upLimitWord:
            i=i+1
        while sh.cell(i,ColumTotal).value !=downLimitWord:
            diferencesDict={
                "Concept":sh.cell(i,columConcept).value,
                "Motive":sh.cell(i,columMotive).value,
                "SubtotalUs":sh.cell(i,ColumSubtotalUs).value,
                "SubtotalBs":sh.cell(i,ColumSubtotalBs).value,
                "TotalBs":sh.cell(i,ColumTotalBs).value,
            }
            
            if diferencesDict["Concept"] not in keywordsFilter:
                diferencesTable.append(diferencesDict)
            i=i+1
        print(pd.DataFrame(diferencesTable))
        return pd.DataFrame(diferencesTable)

def scrapXlsxFile(fileName,distributionType):
    
    scrapyxlsx=scrapTablesExcel(fileName,distributionType)
    if distributionType=="distribuidora":
        if fileName.find("First")!=-1:
            summaryTable=scrapyxlsx.getSummaryTable()
        elif fileName.find("Us")!=-1:
            billTable=scrapyxlsx.getBillstable()
            checkTable=scrapyxlsx.getChecksTable()
            bankTransferTable=scrapyxlsx.getBankTransfersTable()
        elif fileName.find("Bs")!=-1:
            billTable=scrapyxlsx.getBillstable()
            coinsTable=scrapyxlsx.getCoinsTable()
            checkTable=scrapyxlsx.getChecksTable()
            bankTransferTable=scrapyxlsx.getBankTransfersTable()
        else:
            print("Error en el nombre del archivo")
    elif distributionType=="agencia":
        if fileName.find("Us")!=-1:
            billTable=scrapyxlsx.getBillstable()
            voucherTable=scrapyxlsx.get_vouchers_table()
            qrTable=scrapyxlsx.get_qr_table()
        elif fileName.find("Bs")!=-1:
            scrapyxlsx.updateCurrency()
            billTable=scrapyxlsx.getBillstable()
            coinsTable=scrapyxlsx.getCoinsTable()
            voucherTable=scrapyxlsx.get_vouchers_table()
            cuoponTable=scrapyxlsx.get_coupon_table()
            diferencesTable=scrapyxlsx.get_diferences_table()
            
        else:
            print("Error en el nombre del archivo")
def scrapDolarOperationsXls():
    with open(r'src\target\CashClosingInfo.json',"r") as json_file:
        data = json.load(json_file)
    for row in data['data']:
        print(row['Código'])
        for path in row["xlsFilesList"]:
            path=path.replace("descargas","descargasXlsx").replace("xls","xlsx")
            print(path)
            scrapXlsxFile(path,row["Acciones"])
def test_scrapDolarOperationsXls():
    xlsxFilesList=os.listdir(r"descargasXlsx")
    for xlsxFile in xlsxFilesList:
        print(xlsxFile)
        scrapXlsxFile(xlsxFile,"agencia")
        #scrapXlsxFile(r"src\target\descargasXlsx\\"+xlsxFile,"agencia")             
        
#scrapDolarOperationsXls()
test_scrapDolarOperationsXls()
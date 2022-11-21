# playwright 
from playwright.sync_api import sync_playwright
from datetime import datetime
import pandas as pd
from pathlib import Path
import openpyxl
import os
import time
import locale
def init_page():
    # Go to https://www.borrd.com/
    page.goto("http://sgv.grupo-venado.com/venado/login.jsf")
    page.get_by_placeholder("Usuario").click()
    page.get_by_placeholder("Usuario").fill("BOT.ADMINISTRACION.LP")
    page.get_by_placeholder("Contraseña").click()
    page.get_by_placeholder("Contraseña").fill("venadobot")
    page.get_by_role("button", name="Iniciar Sesión").click()
    page.wait_for_load_state()

def goto_bills():
    page.get_by_role("link", name="  Cobranza").click()
    page.get_by_role("link", name=" Cierres de Caja").click()
    page.wait_for_load_state()

def set_day(dExcel,cssDate):
    if cssDate=="input#startDate":
        dates=[x for x in page.query_selector_all("div:nth-child(10) div.datepicker-days tbody td[class='day']")]
    elif cssDate=="input#endDate":
        dates=[x for x in page.query_selector_all("div:nth-child(11) div.datepicker-days tbody td[class='day']")]
    else:
        return
    for d in dates:
        if d.inner_text()==dExcel.strftime("%d"):
            d.click()
            break
def tableCashClosing():
    table=[]
    headersTable=[x.inner_text() for x in page.query_selector_all("table#cashierClosings thead th")]
    rows=page.query_selector_all("table#cashierClosings tbody tr")
    print(len(rows))
    for row in rows:
        if len(row.query_selector_all("a"))==7:
            tipe="agencia"
        elif len(row.query_selector_all("a"))==5:
            tipe="distribuidora"
        else:
            tipe="otro"
        xpathArceoCajaBs="a[data-original-title='Arqueo de Caja Bs. EXCEL']"
        xpathArceoCajaUs="a[data-original-title='Arqueo de Caja $us. EXCEL']"
        xpathFirstExcel="a[data-original-title='Descargar EXCEL']"
        fields=[y.inner_text() for y in row.query_selector_all("td")]
        cashCode=fields[0]
        if tipe=="agencia":
            download_file(f"{cashCode}_arceoCajaBs.xls",xpathArceoCajaBs)
            download_file(f"{cashCode}_arceoCajaUs.xls",xpathArceoCajaUs)
            download_file(f"{cashCode}_firstExcel.xls",xpathFirstExcel)
        elif tipe=="distribuidora":
            download_file(f"{cashCode}_arceoCajaBs.xls",xpathArceoCajaBs)
            download_file(f"{cashCode}_arceoCajaUs.xls",xpathArceoCajaUs)
        else:
            pass 
        
        rowDict={
            headersTable[0]:fields[0],
            headersTable[1]:fields[1],
            headersTable[2]:fields[2],
            headersTable[3]:fields[3],
            headersTable[4]:fields[4],
            headersTable[5]:fields[5],
            headersTable[6]:fields[6],
            headersTable[7]:fields[7],
            headersTable[8]:fields[8],
            headersTable[9]:tipe
        }
        table.append(rowDict)
    return pd.DataFrame(table)
def evaluate_month(monthdate_obj,dExcel,cssDate):
    tday=dExcel.strftime("%B %Y")
    if monthdate_obj.strftime("%B %Y")==tday:
        print("same month")
        set_day(dExcel,cssDate)
        return True
    elif monthdate_obj<dExcel:
        print("next month")
        page.query_selector("div.datepicker-days th.next").click()
        return False
        #monthdate=w.find_element(By.CSS_SELECTOR,"div.datepicker-days th.datepicker-switch").text
    elif monthdate_obj>dExcel:
        print("previous month")
        page.query_selector("div.datepicker-days th.prev").click()
        return False
def found_date(dExcel,cssDate):
    page.query_selector(cssDate).click()
    if cssDate=="input#startDate":
        monthdate=page.query_selector("body > div:nth-child(10) > div.datepicker-days > table > thead > tr:nth-child(1) > th.datepicker-switch").inner_text()
        monthdate=monthdate.replace("Septiembre","Setiembre")
        monthdate_obj=datetime.strptime(monthdate,"%B %Y")
    elif cssDate=="input#endDate":
        monthdate=page.query_selector("body > div:nth-child(11) > div.datepicker-days > table > thead > tr:nth-child(1) > th.datepicker-switch").inner_text()
        monthdate=monthdate.replace("Septiembre","Setiembre")
        print(monthdate)
        monthdate_obj=datetime.strptime(monthdate,"%B %Y")
   
    dateNotfound=True
    while dateNotfound:
        if evaluate_month(monthdate_obj,dExcel,cssDate):
            dateNotfound=False
            print("date found")
        else:
            monthdate=page.query_selector("div.datepicker-days th.datepicker-switch").inner_text()
            monthdate=monthdate.replace("Septiembre","Setiembre")
            monthdate_obj=datetime.strptime(monthdate,"%B %Y")

def set_dates(dinit,dEnd):
    found_date(dinit,"input#startDate")
    time.sleep(1)
    found_date(dEnd,"input#endDate")

    page.get_by_placeholder("Fecha inicial").click()
def in_folder(nameFolder):
    folderParent = os.getcwd()
    #folderParent=Path(folderParent).parent
    folderParent=os.path.join(folderParent,nameFolder)
    return folderParent
def download_file(nameFile,cssSelector):
    with page.expect_download() as download_info:
        page.query_selector(cssSelector).click()
    download = download_info.value
    nameFile=os.path.join(in_folder("descargas"),nameFile)
    download.save_as(nameFile)

def main():
    wb=openpyxl.load_workbook("src\config.xlsx")
    ws=wb["Hoja1"]
    dinit=ws["B2"].value
    dEnd=ws["B3"].value
    locale.setlocale(locale.LC_TIME, '')
    with sync_playwright() as p:
        global browser,context,page
        browser = p.chromium.launch(headless=False)
        context = browser.new_context()
                # Open new page
        page = context.new_page()
        init_page()
        goto_bills()
        found_date(dinit,"input#startDate")
        time.sleep(1)
        found_date(dEnd,"input#endDate")
        df=tableCashClosing()
        print(df)
        page.pause()
if __name__ == "__main__":
    main()